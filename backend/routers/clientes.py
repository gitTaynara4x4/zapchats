# backend/routers/clientes.py
from __future__ import annotations

from datetime import date, datetime, time, timezone, timedelta
from typing import Optional, List, Any, Dict, Set

from fastapi import APIRouter, Depends, HTTPException, Header, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import desc, or_, and_, func
from sqlalchemy.orm import Session, aliased
from sqlalchemy.exc import IntegrityError

import io
import csv
import json
import hashlib
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None  # type: ignore


def tz_sao_paulo():
    if ZoneInfo:
        try:
            return ZoneInfo("America/Sao_Paulo")
        except Exception:
            pass
    return timezone(timedelta(hours=-3))


from backend.database import get_db
from backend import models
from backend.routers.auth import get_current_identity
from backend.utils.entitlements import (
    enforce_quota,
    enforce_feature,
)
from backend.security.instancias import instancias_visiveis
from backend.security.atendimento_acl import (
    VISIBILIDADE_ATENDIMENTOS_PROPRIOS,
    assert_cliente_access,
    get_atendimento_visibility_scope,
)
from backend.services.atendimento_claim_state import (
    ensure_open_atendimento_locked,
    get_open_atendimento_locked,
    join_participant,
)
from backend.migrations.clientes_sequence import sync_clientes_id_sequence
from backend.cache.redis_client import (
    get_json as cache_get_json,
    set_json as cache_set_json,
    delete_prefix as cache_delete_prefix,
    k as cache_k,
)

router = APIRouter(prefix="/clientes", tags=["Clientes"])


class ClienteWriteFields(BaseModel):
    nome: Optional[str] = None
    telefone: Optional[str] = None
    departamento: Optional[str] = None
    departamento_id: Optional[int] = None
    colaborador_id: Optional[int] = None
    sobre_cliente: Optional[str] = None
    cpf_cnpj: Optional[str] = None
    rg: Optional[str] = None
    email: Optional[str] = None
    data_nascimento: Optional[date | datetime] = None
    genero: Optional[str] = None
    cep: Optional[str] = None
    endereco: Optional[str] = None
    numero: Optional[str] = None
    complemento: Optional[str] = None
    bairro: Optional[str] = None
    cidade: Optional[str] = None
    estado: Optional[str] = None
    nome_completo: Optional[str] = None
    website: Optional[str] = None
    descricao: Optional[str] = None
    nome_whatsapp: Optional[str] = None
    status_whatsapp: Optional[str] = None
    is_business: Optional[bool] = None


class PatchClienteProfile(ClienteWriteFields):
    pass


class PostNovoCliente(ClienteWriteFields):
    telefone: str
    instancia_id: Optional[int] = None
    instance_name: Optional[str] = None
    # Marcador explícito para diferenciar o cadastro completo de clientes
    # da criação mínima feita pelo botão "Nova conversa" no atendimento.
    origem_atendimento: bool = False


class BulkColaboradorIn(BaseModel):
    ids: List[int]
    colaborador_id: Optional[int]


class BulkDepartamentoIn(BaseModel):
    ids: List[int]
    departamento_id: Optional[int]


def _id_get(obj: Any, key: str, default: Any = None) -> Any:
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def resolve_empresa_id(
    x_empresa_id: Optional[int] = Header(default=None, alias="X-Empresa-Id"),
    empresa_id_qs: Optional[int] = Query(default=None, alias="empresa_id"),
) -> int:
    emp = x_empresa_id or empresa_id_qs
    if not emp:
        raise HTTPException(
            status_code=400,
            detail="empresa_id é obrigatório (X-Empresa-Id ou ?empresa_id=)",
        )
    return int(emp)


def _get_empresa_or_404(db: Session, empresa_id: int) -> models.Empresa:
    emp = db.query(models.Empresa).filter(models.Empresa.id == int(empresa_id)).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    return emp


def _digits(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())


def _phone_variants(raw_phone: Optional[str]) -> Set[str]:
    """Variações seguras para localizar o mesmo número com ou sem DDI 55."""
    digits = _digits(raw_phone or "")
    if not digits:
        return set()

    variants: Set[str] = {digits}
    national = digits[2:] if digits.startswith("55") and len(digits) in (12, 13) else digits

    if len(national) in (10, 11):
        variants.add(national)
        variants.add(f"55{national}")

    # Compatibilidade com contatos brasileiros antigos salvos sem o nono dígito.
    if len(national) == 11 and national[2:3] == "9":
        old_national = f"{national[:2]}{national[3:]}"
        variants.add(old_national)
        variants.add(f"55{old_national}")
    elif len(national) == 10:
        with_nine = f"{national[:2]}9{national[2:]}"
        variants.add(with_nine)
        variants.add(f"55{with_nine}")

    return {value for value in variants if value}


def _find_cliente_by_phone(
    db: Session,
    *,
    empresa_id: int,
    raw_phone: Optional[str],
) -> Optional[models.Cliente]:
    variants = sorted(_phone_variants(raw_phone), key=len, reverse=True)
    if not variants:
        return None

    national_variants = [value for value in variants if len(value) in (10, 11)]
    clauses = [models.Cliente.telefone_norm.in_(variants)]
    clauses.extend(
        func.right(models.Cliente.telefone_norm, len(value)) == value
        for value in national_variants
    )

    return (
        db.query(models.Cliente)
        .filter(
            models.Cliente.empresa_id == int(empresa_id),
            or_(*clauses),
        )
        .order_by(models.Cliente.id.asc())
        .first()
    )


def _integrity_info(exc: IntegrityError) -> tuple[Optional[str], Optional[str], str]:
    orig = getattr(exc, "orig", None)
    diag = getattr(orig, "diag", None)
    constraint = getattr(diag, "constraint_name", None)
    pgcode = getattr(orig, "pgcode", None)
    first_line = str(orig or exc).splitlines()[0][:500]
    return constraint, pgcode, first_line


def _is_clientes_pkey_error(exc: IntegrityError) -> bool:
    constraint, pgcode, message = _integrity_info(exc)
    return bool(
        constraint == "clientes_pkey"
        or (pgcode == "23505" and "clientes_pkey" in message)
    )


def _log_integrity_error(context: str, exc: IntegrityError) -> None:
    constraint, pgcode, message = _integrity_info(exc)
    print(
        f"[CLIENTES][{context}][INTEGRITY] "
        f"constraint={constraint or '-'} pgcode={pgcode or '-'} error={message}"
    )


def _iso(dt):
    if not dt:
        return None
    if isinstance(dt, datetime):
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.isoformat(timespec="seconds")
    if isinstance(dt, date):
        return dt.isoformat()
    return str(dt)


def _get_perms(identity) -> set[str]:
    return set(_id_get(identity, "permissoes", []) or [])


def _has_any_perm(identity: Any, *wanted: str) -> bool:
    perms = {str(p).strip().lower() for p in _get_perms(identity)}
    return any(str(perm).strip().lower() in perms for perm in wanted)


def _identity_colaborador_id(identity: Any) -> Optional[int]:
    for key in ("colaborador_id", "id_colab", "id_colaborador", "colab_id", "cid"):
        raw = _id_get(identity, key)
        try:
            value = int(raw) if raw is not None else None
        except Exception:
            value = None
        if value and value > 0:
            return value

    sub = str(_id_get(identity, "sub") or "").strip().lower()
    if sub.startswith("colab-"):
        try:
            value = int(sub.split("-", 1)[1])
            return value if value > 0 else None
        except Exception:
            return None

    return None


def _conv_ref_cliente(cliente_id: int, instancia_id: Optional[int]) -> str:
    return f"c:{int(cliente_id)}:{int(instancia_id or 0)}"


def _cliente_conversation_fields(
    *,
    cliente_id: int,
    instancia_id: Optional[int],
) -> Dict[str, Any]:
    conv_key = _conv_ref_cliente(int(cliente_id), instancia_id)
    return {
        "conversation_key": conv_key,
        "conversation_id": conv_key,
        "kind": "c",
        "entity_id": int(cliente_id),
        "cliente_id": int(cliente_id),
    }


def _public_avatar_url(
    *,
    kind: str,
    conversation_id: int,
    raw_avatar_url: Optional[str],
) -> Optional[str]:
    """
    Nunca devolve pps.whatsapp.net direto para o front.

    O banco pode guardar a URL crua da Evolution/WhatsApp,
    mas o navegador deve carregar pelo proxy local:

      /api/atendimento/avatar/{id}?kind=cliente
      /api/atendimento/avatar/{id}?kind=grupo
    """
    if not conversation_id:
        return None

    raw = (raw_avatar_url or "").strip()
    if not raw:
        return None

    if raw.startswith("/api/atendimento/avatar/"):
        return raw

    return f"/api/atendimento/avatar/{int(conversation_id)}?kind={kind}"


def _assert_instancia_acl(identity: Any, db: Session, instancia_id: Optional[int]) -> None:
    vis = instancias_visiveis(identity, db)
    if vis is None:
        return
    if instancia_id is None:
        return
    try:
        iid = int(instancia_id)
    except Exception:
        raise HTTPException(status_code=400, detail="instancia_id inválido")
    if iid not in vis:
        raise HTTPException(status_code=403, detail="Sem acesso a esta instância")


def _resolve_create_instancia(
    db: Session,
    *,
    empresa_id: int,
    identity: Any,
    instancia_id: Optional[int],
    instance_name: Optional[str],
) -> Optional[models.EmpresaInstancia]:
    iid = int(instancia_id) if instancia_id is not None else None
    name = (instance_name or "").strip()

    if iid is None and not name:
        return None

    query = db.query(models.EmpresaInstancia).filter(
        models.EmpresaInstancia.empresa_id == int(empresa_id),
    )

    if iid is not None:
        row = query.filter(models.EmpresaInstancia.id == iid).first()
    else:
        row = query.filter(models.EmpresaInstancia.instance_name == name).first()

    if not row:
        raise HTTPException(status_code=404, detail="WhatsApp selecionado não foi encontrado.")

    _assert_instancia_acl(identity, db, int(row.id))
    return row


def _apply_instancias_filter(identity: Any, db: Session, query):
    vis = instancias_visiveis(identity, db)
    if vis is None:
        return query
    if not vis:
        return query.filter(models.Cliente.id == -1)
    return query.filter(models.Cliente.instancia_id.in_(vis))


def get_empresa_autorizada(
    empresa_id: int = Depends(resolve_empresa_id),
    identity=Depends(get_current_identity),
) -> int:
    emp = _id_get(identity, "empresa_id")
    if emp is not None and int(emp) != int(empresa_id):
        raise HTTPException(
            status_code=403,
            detail="Empresa inválida para este usuário",
        )
    return int(empresa_id)


COLUMNS = ["nome", "telefone", "departamento", "sobre_cliente"]


def _clean_text(x: str | None) -> str:
    return (x or "").replace("\r", "").strip()


def _nullable_text(value: Any) -> Optional[str]:
    cleaned = str(value or "").replace("\r", "").strip()
    return cleaned or None


def _birth_datetime(value: date | datetime | None) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    return datetime.combine(value, time.min)


def _validate_colaborador_id(
    db: Session,
    *,
    empresa_id: int,
    colaborador_id: Optional[int],
) -> Optional[int]:
    if colaborador_id is None:
        return None
    row = (
        db.query(models.Colaborador.id)
        .filter(
            models.Colaborador.id == int(colaborador_id),
            models.Colaborador.empresa_id == int(empresa_id),
        )
        .first()
    )
    if not row:
        raise HTTPException(
            status_code=400,
            detail="colaborador_id inválido para esta empresa",
        )
    return int(colaborador_id)


def _resolve_departamento(
    db: Session,
    *,
    empresa_id: int,
    departamento_id: Optional[int],
    departamento: Optional[str],
) -> tuple[Optional[int], Optional[str]]:
    if departamento_id is not None:
        row = (
            db.query(models.Departamento.id, models.Departamento.nome)
            .filter(
                models.Departamento.id == int(departamento_id),
                models.Departamento.empresa_id == int(empresa_id),
            )
            .first()
        )
        if not row:
            raise HTTPException(
                status_code=400,
                detail="departamento_id inválido para esta empresa",
            )
        return int(row.id), _nullable_text(row.nome)

    nome = _nullable_text(departamento)
    if not nome:
        return None, None

    row = (
        db.query(models.Departamento.id, models.Departamento.nome)
        .filter(
            models.Departamento.empresa_id == int(empresa_id),
            func.lower(func.trim(models.Departamento.nome)) == nome.lower(),
        )
        .first()
    )
    if row:
        return int(row.id), _nullable_text(row.nome)

    # Mantém compatibilidade com cadastros antigos que usavam texto livre,
    # mas não associa um ID incorreto a filas/triagem.
    return None, nome


def _normalize_unique_phone(
    db: Session,
    *,
    empresa_id: int,
    raw_phone: Optional[str],
    exclude_cliente_id: Optional[int] = None,
) -> str:
    telefone = _digits(raw_phone or "")
    if len(telefone) < 8:
        raise HTTPException(status_code=400, detail="Telefone inválido")

    query = db.query(models.Cliente.id).filter(
        models.Cliente.empresa_id == int(empresa_id),
        models.Cliente.telefone_norm == telefone,
    )
    if exclude_cliente_id is not None:
        query = query.filter(models.Cliente.id != int(exclude_cliente_id))
    if query.first():
        raise HTTPException(
            status_code=400,
            detail="Já existe um cliente com esse telefone.",
        )
    return telefone


CLIENTE_TEXT_FIELDS = (
    "sobre_cliente",
    "cpf_cnpj",
    "rg",
    "email",
    "genero",
    "cep",
    "endereco",
    "numero",
    "complemento",
    "bairro",
    "cidade",
    "estado",
    "nome_completo",
    "website",
    "descricao",
    "nome_whatsapp",
    "status_whatsapp",
)


def _apply_cliente_payload(
    db: Session,
    *,
    cliente: models.Cliente,
    payload: ClienteWriteFields,
    empresa_id: int,
    creating: bool = False,
) -> None:
    fields = set(payload.model_fields_set)

    if creating or "nome" in fields:
        cliente.nome = _nullable_text(payload.nome) or "Cliente"

    if "telefone" in fields:
        telefone = _normalize_unique_phone(
            db,
            empresa_id=empresa_id,
            raw_phone=payload.telefone,
            exclude_cliente_id=(None if creating else int(cliente.id)),
        )
        # telefone_norm é GENERATED ALWAYS no PostgreSQL e será
        # recalculado automaticamente a partir de telefone.
        cliente.telefone = telefone

    if "colaborador_id" in fields:
        cliente.colaborador_id = _validate_colaborador_id(
            db,
            empresa_id=empresa_id,
            colaborador_id=payload.colaborador_id,
        )

    if "departamento_id" in fields or "departamento" in fields:
        dep_id, dep_nome = _resolve_departamento(
            db,
            empresa_id=empresa_id,
            departamento_id=payload.departamento_id,
            departamento=payload.departamento,
        )
        cliente.departamento_id = dep_id
        cliente.departamento = dep_nome

    for field in CLIENTE_TEXT_FIELDS:
        if field in fields:
            setattr(cliente, field, _nullable_text(getattr(payload, field)))

    if "data_nascimento" in fields:
        cliente.data_nascimento = _birth_datetime(payload.data_nascimento)

    if "is_business" in fields:
        cliente.is_business = bool(payload.is_business)


def _auto_sep(sample: str) -> str:
    return ";" if sample.count(";") >= sample.count(",") else ","


def _xlsx_rows_to_dicts(ws):
    rows = list(ws.rows)
    if not rows:
        return []
    head = [str((c.value or "")).strip().lower() for c in rows[0]]
    start = 1 if all(h in COLUMNS for h in head[:4]) else 0
    out = []
    for r in rows[start:]:
        vals = [str(c.value or "").strip() for c in r[:4]] + ["", "", "", ""]
        out.append(
            {
                "nome": vals[0],
                "telefone": vals[1],
                "departamento": vals[2],
                "sobre_cliente": vals[3],
            }
        )
    return out


@router.get("")
def listar_clientes(
    empresa_id: int = Depends(get_empresa_autorizada),
    q: Optional[str] = Query(None),
    departamento: Optional[str] = Query(None),
    departamento_id: Optional[int] = Query(None),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    instancia_id: Optional[int] = Query(None),
    colaborador_id: Optional[int] = Query(None),
    limit: int = Query(20, ge=1, le=200),
    offset: int = Query(0, ge=0),
    include_total: bool = Query(
        True,
        description="Quando false, evita COUNT(*) e prioriza a primeira renderização.",
    ),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.ver" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para ver clientes (clientes.ver).",
        )

    if instancia_id is not None:
        _assert_instancia_acl(identity, db, instancia_id)

    key_src = json.dumps(
        {
            "v": "clientes-conversation-key-v3-avatar-proxy",
            "emp": empresa_id,
            "q": (q or "").strip(),
            "departamento": (departamento or "").strip(),
            "departamento_id": departamento_id,
            "di": str(data_inicio) if data_inicio else "",
            "df": str(data_fim) if data_fim else "",
            "inst": instancia_id,
            "colab": colaborador_id,
            "limit": limit,
            "offset": offset,
            "include_total": bool(include_total),
            "acl": instancias_visiveis(identity, db),
        },
        sort_keys=True,
        ensure_ascii=False,
    )
    hashed = hashlib.sha256(key_src.encode("utf-8")).hexdigest()
    cache_key = cache_k("clientes", "list", str(empresa_id), hashed)

    cached = cache_get_json(cache_key)
    if cached:
        return cached

    Colab = aliased(models.Colaborador)
    Inst = aliased(models.EmpresaInstancia)

    base = (
        db.query(
            models.Cliente.id,
            models.Cliente.nome,
            models.Cliente.nome_whatsapp,
            models.Cliente.telefone,
            models.Cliente.avatar_url,
            models.Cliente.departamento,
            models.Cliente.departamento_id,
            models.Cliente.sobre_cliente.label("sobre"),
            models.Cliente.timestamp.label("data_cadastro"),
            models.Cliente.timestamp.label("timestamp"),
            models.Cliente.colaborador_id,
            Colab.nome.label("colaborador_nome"),
            models.Cliente.instancia_id,
            Inst.instance_name.label("instance_name"),
        )
        .outerjoin(
            Colab,
            and_(
                Colab.id == models.Cliente.colaborador_id,
                Colab.empresa_id == models.Cliente.empresa_id,
            ),
        )
        .outerjoin(
            Inst,
            and_(
                Inst.id == models.Cliente.instancia_id,
                Inst.empresa_id == models.Cliente.empresa_id,
            ),
        )
        .filter(models.Cliente.empresa_id == empresa_id)
    )

    base = _apply_instancias_filter(identity, db, base)

    if instancia_id is not None:
        base = base.filter(models.Cliente.instancia_id == instancia_id)

    if q:
        like = f"%{q.strip()}%"
        dq = _digits(q)
        conds = [
            models.Cliente.nome.ilike(like),
            models.Cliente.nome_whatsapp.ilike(like),
            models.Cliente.telefone.ilike(like),
        ]
        if dq:
            conds.append(models.Cliente.telefone_norm.ilike(f"%{dq}%"))
        base = base.filter(or_(*conds))

    if departamento_id is not None:
        base = base.filter(models.Cliente.departamento_id == int(departamento_id))
    elif departamento:
        base = base.filter(func.lower(models.Cliente.departamento) == departamento.strip().lower())

    if colaborador_id is not None:
        if colaborador_id == 0:
            base = base.filter(models.Cliente.colaborador_id.is_(None))
        else:
            base = base.filter(models.Cliente.colaborador_id == colaborador_id)

    tz_sp = tz_sao_paulo()
    if data_inicio:
        ini_local = datetime.combine(data_inicio, time(0, 0, 0, tzinfo=tz_sp))
        ini_utc = ini_local.astimezone(timezone.utc)
        base = base.filter(models.Cliente.timestamp >= ini_utc)
    if data_fim:
        fim_local = datetime.combine(data_fim, time(23, 59, 59, 999_999, tzinfo=tz_sp))
        fim_utc = fim_local.astimezone(timezone.utc)
        base = base.filter(models.Cliente.timestamp <= fim_utc)

    ordered = base.order_by(desc(models.Cliente.timestamp), desc(models.Cliente.id))

    if include_total:
        total: Optional[int] = int(base.order_by(None).count())
        rows = ordered.offset(offset).limit(limit).all()
        has_more = (offset + len(rows)) < total
    else:
        # A tela precisa dos registros antes da contagem exata. Buscar um item
        # extra permite descobrir se existe próxima página sem executar COUNT(*),
        # que pode ser caro em empresas com muitos clientes e filtros.
        rows_plus_one = ordered.offset(offset).limit(limit + 1).all()
        has_more = len(rows_plus_one) > limit
        rows = rows_plus_one[:limit]
        total = None

    items = []
    for rrow in rows:
        fields = _cliente_conversation_fields(
            cliente_id=int(rrow.id),
            instancia_id=rrow.instancia_id,
        )

        items.append(
            {
                "id": rrow.id,
                **fields,
                "nome": rrow.nome,
                "nome_whatsapp": rrow.nome_whatsapp,
                "telefone": rrow.telefone,
                "avatar_url": _public_avatar_url(
                    kind="cliente",
                    conversation_id=int(rrow.id),
                    raw_avatar_url=rrow.avatar_url,
                ),
                "departamento": rrow.departamento,
                "departamento_id": rrow.departamento_id,
                "sobre": (rrow.sobre or None),
                "data_cadastro": _iso(getattr(rrow, "data_cadastro", None)),
                "timestamp": _iso(getattr(rrow, "timestamp", None)),
                "colaborador_id": rrow.colaborador_id,
                "colaborador_nome": rrow.colaborador_nome,
                "instancia_id": rrow.instancia_id,
                "instance_name": rrow.instance_name,
                "is_group": False,
            }
        )

    out = {
        "items": items,
        "total": total,
        "has_more": has_more,
        "next_offset": (offset + len(items)) if has_more else None,
        "limit": limit,
        "offset": offset,
    }

    cache_set_json(cache_key, out)
    return out


@router.get("/search")
def buscar_clientes_leve(
    empresa_id: int = Depends(get_empresa_autorizada),
    q: str = Query(..., min_length=1),
    limit: int = Query(8, ge=1, le=20),
    instancia_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.ver" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para buscar clientes.",
        )

    termo = (q or "").strip()
    if not termo:
        return {"items": []}

    if instancia_id is not None:
        _assert_instancia_acl(identity, db, instancia_id)

    like = f"%{termo}%"
    dq = _digits(termo)

    Inst = aliased(models.EmpresaInstancia)

    query = (
        db.query(
            models.Cliente.id,
            models.Cliente.nome,
            models.Cliente.nome_whatsapp,
            models.Cliente.telefone,
            models.Cliente.avatar_url,
            models.Cliente.instancia_id,
            Inst.instance_name.label("instance_name"),
        )
        .outerjoin(
            Inst,
            and_(
                Inst.id == models.Cliente.instancia_id,
                Inst.empresa_id == models.Cliente.empresa_id,
            ),
        )
        .filter(models.Cliente.empresa_id == empresa_id)
    )

    query = _apply_instancias_filter(identity, db, query)

    if instancia_id is not None:
        query = query.filter(models.Cliente.instancia_id == instancia_id)

    conds = [
        models.Cliente.nome.ilike(like),
        models.Cliente.nome_whatsapp.ilike(like),
        models.Cliente.telefone.ilike(like),
    ]
    if dq:
        conds.append(models.Cliente.telefone_norm.ilike(f"%{dq}%"))

    rows = (
        query
        .filter(or_(*conds))
        .order_by(desc(models.Cliente.timestamp), desc(models.Cliente.id))
        .limit(limit)
        .all()
    )

    items = []
    for r in rows:
        nome = (r.nome_whatsapp or r.nome or "Cliente").strip()
        fields = _cliente_conversation_fields(
            cliente_id=int(r.id),
            instancia_id=r.instancia_id,
        )

        items.append({
            "id": r.id,
            **fields,
            "nome": nome,
            "nome_whatsapp": r.nome_whatsapp,
            "telefone": r.telefone,
            "avatar_url": _public_avatar_url(
                kind="cliente",
                conversation_id=int(r.id),
                raw_avatar_url=r.avatar_url,
            ),
            "instancia_id": r.instancia_id,
            "instance_name": r.instance_name,
            "is_group": False,
        })

    return {"items": items}


@router.get("/departamentos-resumo")
def listar_departamentos_resumo(
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    """Lista mínima para filtros e ficha de clientes.

    A rota completa de Departamentos também calcula membros, instâncias e outros
    dados por departamento. A tela de Clientes precisa apenas de id e nome.
    """
    perms = _get_perms(identity)
    if "clientes.ver" not in perms and "clientes.editar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para listar departamentos para clientes.",
        )

    rows = (
        db.query(models.Departamento.id, models.Departamento.nome)
        .filter(models.Departamento.empresa_id == int(empresa_id))
        .order_by(func.lower(models.Departamento.nome).asc())
        .all()
    )

    return {
        "items": [
            {"id": int(row.id), "nome": row.nome or "Departamento"}
            for row in rows
        ]
    }


@router.get("/colaboradores")
def listar_colaboradores(
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.ver" not in perms and "clientes.editar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para listar colaboradores para clientes.",
        )

    q = (
        db.query(models.Colaborador.id, models.Colaborador.nome, models.Colaborador.email)
        .filter(models.Colaborador.empresa_id == empresa_id)
        .order_by(models.Colaborador.nome.asc())
    )
    items = [{"id": r.id, "nome": r.nome, "email": r.email} for r in q.all()]
    return {"items": items}


@router.get("/export")
def exportar_clientes(
    fmt: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    ids: Optional[str] = Query(None, description="IDs separados por vírgula (ex: 1,2,3)"),
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.importar_exportar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para exportar clientes (clientes.importar_exportar).",
        )

    empresa = _get_empresa_or_404(db, empresa_id)

    enforce_feature(
        empresa,
        "feature_export",
        message="Seu plano não permite exportação de clientes ou está vencido. Renove para continuar.",
    )

    q = db.query(models.Cliente).filter(models.Cliente.empresa_id == empresa_id)
    q = _apply_instancias_filter(identity, db, q)

    ids_set: set[int] = set()
    if ids:
        for x in ids.split(","):
            x = (x or "").strip()
            if x.isdigit():
                ids_set.add(int(x))
    if ids_set:
        q = q.filter(models.Cliente.id.in_(list(ids_set)))

    rows: list[models.Cliente] = q.order_by(models.Cliente.nome.asc()).all()

    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";")
        w.writerow(["nome", "telefone", "departamento", "sobre_cliente"])
        for r in rows:
            w.writerow(
                [
                    _clean_text(r.nome),
                    _clean_text(r.telefone),
                    _clean_text(r.departamento),
                    _clean_text(r.sobre_cliente),
                ]
            )
        data = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="clientes.csv"'},
        )

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        ws.append(["nome", "telefone", "departamento", "sobre_cliente"])
        for r in rows:
            ws.append([r.nome or "", r.telefone or "", r.departamento or "", r.sobre_cliente or ""])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="clientes.xlsx"'},
        )

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    left, top = 40, height - 50
    c.setFont("Helvetica-Bold", 12)
    c.drawString(left, top, "Relatório de Clientes")
    c.setFont("Helvetica", 9)
    y = top - 24
    c.drawString(left, y, "Nome")
    c.drawString(left + 220, y, "Telefone")
    c.drawString(left + 360, y, "Departamento")
    y -= 16
    c.line(left, y + 12, width - 40, y + 12)

    for r in rows:
        if y < 50:
            c.showPage()
            y = height - 60
            c.setFont("Helvetica", 9)
        c.drawString(left, y, (r.nome or "")[:32])
        c.drawString(left + 220, y, (r.telefone or "")[:20])
        c.drawString(left + 360, y, (r.departamento or "")[:18])
        y -= 14

    c.showPage()
    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="clientes.pdf"'},
    )


@router.get("/modelo-importacao")
def baixar_modelo_importacao(
    fmt: str = Query("csv", pattern="^(csv|xlsx|txt)$"),
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.importar_exportar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para baixar o modelo de importação.",
        )

    empresa = _get_empresa_or_404(db, empresa_id)
    enforce_feature(
        empresa,
        "feature_import",
        message="Seu plano não permite importação de clientes ou está vencido.",
    )

    headers = ["nome", "telefone", "departamento", "sobre_cliente"]
    example = ["Maria Souza", "5511999999999", "Comercial", "Cliente preferencial"]

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        ws.append(headers)
        ws.append(example)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="modelo-clientes.xlsx"'},
        )

    if fmt == "txt":
        data = "5511999999999\n5511888888888\n".encode("utf-8")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="modelo-clientes.txt"'},
        )

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    writer.writerow(example)
    data = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="modelo-clientes.csv"'},
    )


@router.get("/{cliente_id}")
def obter_cliente(
    cliente_id: int,
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.ver" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para ver detalhes de clientes.",
        )

    Colab = aliased(models.Colaborador)
    Inst = aliased(models.EmpresaInstancia)

    r = (
        db.query(
            models.Cliente.id,
            models.Cliente.nome,
            models.Cliente.nome_whatsapp,
            models.Cliente.telefone,
            models.Cliente.avatar_url,
            models.Cliente.departamento,
            models.Cliente.departamento_id,
            models.Cliente.sobre_cliente.label("sobre"),
            models.Cliente.timestamp.label("data_cadastro"),
            models.Cliente.timestamp.label("timestamp"),
            models.Cliente.colaborador_id,
            Colab.nome.label("colaborador_nome"),
            models.Cliente.instancia_id,
            Inst.instance_name.label("instance_name"),
            models.Cliente.cpf_cnpj,
            models.Cliente.rg,
            models.Cliente.email,
            models.Cliente.data_nascimento,
            models.Cliente.genero,
            models.Cliente.cep,
            models.Cliente.endereco,
            models.Cliente.numero,
            models.Cliente.complemento,
            models.Cliente.bairro,
            models.Cliente.cidade,
            models.Cliente.estado,
            models.Cliente.nome_completo,
            models.Cliente.website,
            models.Cliente.descricao,
            models.Cliente.is_business,
            models.Cliente.status_whatsapp,
        )
        .outerjoin(
            Colab,
            and_(
                Colab.id == models.Cliente.colaborador_id,
                Colab.empresa_id == models.Cliente.empresa_id,
            ),
        )
        .outerjoin(
            Inst,
            and_(
                Inst.id == models.Cliente.instancia_id,
                Inst.empresa_id == models.Cliente.empresa_id,
            ),
        )
        .filter(models.Cliente.id == cliente_id, models.Cliente.empresa_id == empresa_id)
        .first()
    )
    if not r:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    _assert_instancia_acl(identity, db, r.instancia_id)

    fields = _cliente_conversation_fields(
        cliente_id=int(r.id),
        instancia_id=r.instancia_id,
    )

    return {
        "id": r.id,
        **fields,
        "nome": r.nome,
        "nome_whatsapp": r.nome_whatsapp,
        "telefone": r.telefone,
        "avatar_url": _public_avatar_url(
            kind="cliente",
            conversation_id=int(r.id),
            raw_avatar_url=r.avatar_url,
        ),
        "departamento": r.departamento,
        "departamento_id": r.departamento_id,
        "sobre": r.sobre or None,
        "data_cadastro": _iso(getattr(r, "data_cadastro", None)),
        "timestamp": _iso(getattr(r, "timestamp", None)),
        "colaborador_id": r.colaborador_id,
        "colaborador_nome": r.colaborador_nome,
        "instancia_id": r.instancia_id,
        "instance_name": r.instance_name,
        "is_group": False,
        "cpf_cnpj": r.cpf_cnpj,
        "rg": r.rg,
        "email": r.email,
        "data_nascimento": _iso(r.data_nascimento),
        "genero": r.genero,
        "cep": r.cep,
        "endereco": r.endereco,
        "numero": r.numero,
        "complemento": r.complemento,
        "bairro": r.bairro,
        "cidade": r.cidade,
        "estado": r.estado,
        "nome_completo": r.nome_completo,
        "website": r.website,
        "descricao": r.descricao,
        "is_business": r.is_business,
        "status_whatsapp": r.status_whatsapp,
    }


@router.patch("/{cliente_id}", include_in_schema=False)
@router.patch("/{cliente_id}/profile")
def patch_cliente_profile(
    cliente_id: int,
    payload: PatchClienteProfile,
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.editar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para editar clientes (clientes.editar).",
        )

    c = (
        db.query(models.Cliente)
        .filter(models.Cliente.id == cliente_id, models.Cliente.empresa_id == empresa_id)
        .first()
    )
    if not c:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    _assert_instancia_acl(identity, db, getattr(c, "instancia_id", None))

    _apply_cliente_payload(
        db,
        cliente=c,
        payload=payload,
        empresa_id=empresa_id,
        creating=False,
    )

    db.add(c)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Já existe um cliente com esse telefone.",
        ) from exc

    cache_delete_prefix(cache_k("clientes", "list", str(empresa_id)))
    return {"ok": True, "id": int(c.id)}


@router.post("/novo")
def criar_cliente(
    body: PostNovoCliente,
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    # Cadastro manual de cliente deve respeitar clientes.criar em qualquer tela.
    # `origem_atendimento` informa apenas de onde veio a ação; ele não concede
    # permissão extra. A criação automática de contatos por mensagens recebidas
    # acontece no fluxo interno do atendimento e não passa por este endpoint.
    pode_cadastrar_cliente = _has_any_perm(
        identity,
        "clientes.criar",
        "clientes.gerenciar",
        "admin",
        "root",
    )

    if not pode_cadastrar_cliente:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para criar clientes (clientes.criar).",
        )

    empresa = _get_empresa_or_404(db, empresa_id)
    instancia = _resolve_create_instancia(
        db,
        empresa_id=empresa_id,
        identity=identity,
        instancia_id=body.instancia_id,
        instance_name=body.instance_name,
    )

    visibility_scope = get_atendimento_visibility_scope(
        db,
        identity=identity,
        empresa_id=int(empresa_id),
    )
    current_colab_id = _identity_colaborador_id(identity)
    own_conversations_mode = (
        bool(body.origem_atendimento)
        and visibility_scope == VISIBILIDADE_ATENDIMENTOS_PROPRIOS
        and current_colab_id is not None
    )

    tel = _digits(body.telefone)
    if not tel:
        raise HTTPException(status_code=400, detail="Telefone inválido")

    dup = _find_cliente_by_phone(
        db,
        empresa_id=empresa_id,
        raw_phone=tel,
    )
    if dup:
        selected_instancia_id = (
            int(instancia.id)
            if instancia is not None
            else getattr(dup, "instancia_id", None)
        )

        if bool(body.origem_atendimento) and current_colab_id is not None:
            # Reutiliza o contato existente sem duplicar cadastro. A ACL mantém
            # instância/departamento protegidos; a posse individual não bloqueia
            # a entrada porque o colaborador será adicionado como participante.
            assert_cliente_access(
                db,
                identity=identity,
                empresa_id=int(empresa_id),
                cliente_id=int(dup.id),
                instancia_id=(
                    int(selected_instancia_id)
                    if selected_instancia_id is not None
                    else None
                ),
                allow_unassigned_department=True,
                allow_unowned_if_no_history=True,
                allow_join_existing=True,
            )

            changed = False
            if own_conversations_mode and getattr(dup, "colaborador_id", None) is None:
                dup.colaborador_id = int(current_colab_id)
                changed = True

            if instancia is not None and getattr(dup, "instancia_id", None) is None:
                dup.instancia_id = int(instancia.id)
                selected_instancia_id = int(instancia.id)
                changed = True

            # "Nova conversa" com um contato já em atendimento não toma a
            # responsabilidade de ninguém: o colaborador entra como participante.
            if selected_instancia_id is not None:
                open_atd = get_open_atendimento_locked(
                    db,
                    empresa_id=int(empresa_id),
                    cliente_id=int(dup.id),
                    instancia_id=int(selected_instancia_id),
                )
                if open_atd is None:
                    open_atd = ensure_open_atendimento_locked(
                        db,
                        empresa_id=int(empresa_id),
                        cliente_id=int(dup.id),
                        instancia_id=int(selected_instancia_id),
                        departamento_id=getattr(dup, "departamento_id", None),
                    )
                if open_atd is not None:
                    join_participant(
                        db,
                        atendimento=open_atd,
                        colaborador_id=int(current_colab_id),
                    )
                    changed = True

            if changed:
                db.add(dup)
                db.commit()
                db.refresh(dup)

        fields = _cliente_conversation_fields(
            cliente_id=int(dup.id),
            instancia_id=selected_instancia_id,
        )
        return {
            "id": dup.id,
            **fields,
            "instancia_id": selected_instancia_id,
            "instance_name": getattr(instancia, "instance_name", None),
            "exists": True,
        }

    current_contacts = (
        db.query(func.count(models.Cliente.id))
        .filter(models.Cliente.empresa_id == empresa_id)
        .scalar()
        or 0
    )
    enforce_quota(
        empresa,
        "contacts_max",
        current_contacts,
        delta=1,
        message="Seu plano está vencido ou no limite. Renove para cadastrar novos clientes.",
    )

    c = models.Cliente(
        empresa_id=empresa_id,
        instancia_id=(int(instancia.id) if instancia is not None else None),
        timestamp=datetime.now(timezone.utc),
    )
    _apply_cliente_payload(
        db,
        cliente=c,
        payload=body,
        empresa_id=empresa_id,
        creating=True,
    )

    # No modo "somente minhas conversas", o contato criado no atendimento
    # já nasce vinculado ao colaborador. Assim ele aparece imediatamente e não
    # pode ser aberto por outro atendente com o mesmo escopo.
    if own_conversations_mode:
        c.colaborador_id = int(current_colab_id)

    db.add(c)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        _log_integrity_error("CRIAR", exc)

        dup2 = _find_cliente_by_phone(
            db,
            empresa_id=empresa_id,
            raw_phone=tel,
        )
        if dup2:
            # Corrida de criação: outro request cadastrou o telefone entre a
            # consulta e o INSERT. Aplica a mesma regra de participação usada
            # no caminho normal de contato duplicado.
            selected_instancia_id = (
                int(instancia.id)
                if instancia is not None
                else getattr(dup2, "instancia_id", None)
            )
            if bool(body.origem_atendimento) and current_colab_id is not None:
                assert_cliente_access(
                    db,
                    identity=identity,
                    empresa_id=int(empresa_id),
                    cliente_id=int(dup2.id),
                    instancia_id=(
                        int(selected_instancia_id)
                        if selected_instancia_id is not None
                        else None
                    ),
                    allow_unassigned_department=True,
                    allow_unowned_if_no_history=True,
                    allow_join_existing=True,
                )
                if selected_instancia_id is not None:
                    open_atd = get_open_atendimento_locked(
                        db,
                        empresa_id=int(empresa_id),
                        cliente_id=int(dup2.id),
                        instancia_id=int(selected_instancia_id),
                    )
                    if open_atd is None:
                        open_atd = ensure_open_atendimento_locked(
                            db,
                            empresa_id=int(empresa_id),
                            cliente_id=int(dup2.id),
                            instancia_id=int(selected_instancia_id),
                            departamento_id=getattr(dup2, "departamento_id", None),
                        )
                    if open_atd is not None:
                        join_participant(
                            db,
                            atendimento=open_atd,
                            colaborador_id=int(current_colab_id),
                        )
                        db.commit()

            fields = _cliente_conversation_fields(
                cliente_id=int(dup2.id),
                instancia_id=selected_instancia_id,
            )
            return {
                "id": dup2.id,
                **fields,
                "instancia_id": selected_instancia_id,
                "exists": True,
            }

        # Bancos restaurados/importados podem ficar com clientes_id_seq atrás
        # do maior clientes.id. Nesse caso o INSERT recebe um ID já ocupado.
        if _is_clientes_pkey_error(exc):
            try:
                sync_clientes_id_sequence(db)
                db.commit()

                retry_c = models.Cliente(
                    empresa_id=empresa_id,
                    instancia_id=(int(instancia.id) if instancia is not None else None),
                    timestamp=datetime.now(timezone.utc),
                )
                _apply_cliente_payload(
                    db,
                    cliente=retry_c,
                    payload=body,
                    empresa_id=empresa_id,
                    creating=True,
                )
                if own_conversations_mode:
                    retry_c.colaborador_id = int(current_colab_id)
                db.add(retry_c)
                db.commit()
                c = retry_c
            except IntegrityError as retry_exc:
                db.rollback()
                _log_integrity_error("CRIAR_RETRY", retry_exc)
                raise HTTPException(
                    status_code=409,
                    detail="Não foi possível gerar um novo código para o contato.",
                ) from retry_exc
            except Exception as retry_exc:
                db.rollback()
                print(f"[CLIENTES][CRIAR_RETRY] error={retry_exc}")
                raise HTTPException(
                    status_code=500,
                    detail="Falha ao corrigir a sequência de clientes.",
                ) from retry_exc
        else:
            constraint, pgcode, _ = _integrity_info(exc)
            if pgcode == "23503":
                raise HTTPException(
                    status_code=409,
                    detail="O WhatsApp ou vínculo selecionado não existe mais. Atualize a página e tente novamente.",
                ) from exc
            if pgcode == "23502":
                raise HTTPException(
                    status_code=400,
                    detail="O banco exige um campo obrigatório que não foi preenchido.",
                ) from exc

            raise HTTPException(
                status_code=409,
                detail=f"Não foi possível criar o contato por uma regra do banco ({constraint or 'integridade'}).",
            ) from exc

    db.refresh(c)
    cache_delete_prefix(cache_k("clientes", "list", str(empresa_id)))

    fields = _cliente_conversation_fields(
        cliente_id=int(c.id),
        instancia_id=getattr(c, "instancia_id", None),
    )
    return {
        "id": c.id,
        **fields,
        "instancia_id": getattr(c, "instancia_id", None),
        "instance_name": getattr(instancia, "instance_name", None),
        "created": True,
    }


@router.post("/import")
def importar_clientes(
    arquivo: UploadFile = File(...),
    sobrescrever: bool = Query(False, description="Atualiza dados se telefone já existir"),
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.importar_exportar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para importar clientes (clientes.importar_exportar).",
        )

    empresa = _get_empresa_or_404(db, empresa_id)

    enforce_feature(
        empresa,
        "feature_import",
        message="Seu plano não permite importação de clientes ou está vencido. Renove para continuar.",
    )

    name = (arquivo.filename or "").lower()
    content = arquivo.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Arquivo muito grande. O limite é 10 MB.")

    parsed_rows: List[Dict[str, str]] = []
    if name.endswith(".xlsx"):
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            parsed_rows = _xlsx_rows_to_dicts(ws)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Arquivo XLSX inválido ou corrompido.") from exc
    elif name.endswith(".csv") or name.endswith(".txt"):
        data = content.decode("utf-8-sig", errors="ignore")
        sep = _auto_sep(data)
        reader = csv.reader(io.StringIO(data), delimiter=sep)
        rows = list(reader)
        if not rows:
            raise HTTPException(status_code=400, detail="Arquivo sem linhas.")

        head = [(_clean_text(x)).lower() for x in rows[0]]
        start = 1 if all(h in COLUMNS for h in head[:4]) else 0

        for r in rows[start:]:
            if not r:
                continue
            if len(r) == 1:
                parsed_rows.append({"nome": "", "telefone": r[0], "departamento": "", "sobre_cliente": ""})
            else:
                v = [(r[i] if i < len(r) else "").strip() for i in range(4)]
                parsed_rows.append({"nome": v[0], "telefone": v[1], "departamento": v[2], "sobre_cliente": v[3]})
    else:
        raise HTTPException(status_code=400, detail="Formato não suportado. Use CSV, XLSX ou TXT.")

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="Arquivo sem clientes para importar.")
    if len(parsed_rows) > 50_000:
        raise HTTPException(status_code=413, detail="O arquivo ultrapassa o limite de 50.000 linhas.")

    phones: List[str] = []
    seen: Set[str] = set()
    for row in parsed_rows:
        tel = _digits(row.get("telefone") or "")
        if not tel:
            continue
        if tel not in seen:
            seen.add(tel)
            phones.append(tel)

    existing: Set[str] = set()
    if phones:
        rows_exist = (
            db.query(models.Cliente.telefone_norm)
            .filter(models.Cliente.empresa_id == empresa_id, models.Cliente.telefone_norm.in_(phones))
            .all()
        )
        existing = {str(r[0]) for r in rows_exist if r and r[0]}

    to_insert = len([p for p in phones if p not in existing])

    if to_insert > 0:
        current_contacts = (
            db.query(func.count(models.Cliente.id))
            .filter(models.Cliente.empresa_id == empresa_id)
            .scalar()
            or 0
        )
        enforce_quota(
            empresa,
            "contacts_max",
            current_contacts,
            delta=to_insert,
            message="Seu plano está vencido ou no limite. Renove para importar novos clientes.",
        )

    inseridos = atualizados = ignorados = 0
    processados_no_arquivo: Set[str] = set()

    departamentos = (
        db.query(models.Departamento.id, models.Departamento.nome)
        .filter(models.Departamento.empresa_id == empresa_id)
        .all()
    )
    departamentos_por_nome = {
        str(row.nome or "").strip().lower(): (int(row.id), _nullable_text(row.nome))
        for row in departamentos
        if str(row.nome or "").strip()
    }

    def departamento_values(raw: Optional[str]) -> tuple[Optional[int], Optional[str]]:
        nome = _nullable_text(raw)
        if not nome:
            return None, None
        found = departamentos_por_nome.get(nome.lower())
        if found:
            return found
        return None, nome

    def upsert(nome, telefone, departamento, sobre):
        nonlocal inseridos, atualizados, ignorados
        tel = _digits(telefone or "")
        if len(tel) < 8 or tel in processados_no_arquivo:
            ignorados += 1
            return
        processados_no_arquivo.add(tel)

        dep_id, dep_nome = departamento_values(departamento)
        cli = (
            db.query(models.Cliente)
            .filter(
                models.Cliente.empresa_id == empresa_id,
                models.Cliente.telefone_norm == tel,
            )
            .first()
        )
        if cli:
            _assert_instancia_acl(identity, db, getattr(cli, "instancia_id", None))
            if not sobrescrever:
                ignorados += 1
                return
            nome_limpo = _nullable_text(nome)
            if nome_limpo:
                cli.nome = nome_limpo
            cli.departamento_id = dep_id
            cli.departamento = dep_nome
            cli.sobre_cliente = _nullable_text(sobre)
            db.add(cli)
            atualizados += 1
            return

        novo = models.Cliente(
            empresa_id=empresa_id,
            nome=_nullable_text(nome) or "Cliente",
            telefone=tel,
            departamento_id=dep_id,
            departamento=dep_nome,
            sobre_cliente=_nullable_text(sobre),
            timestamp=datetime.now(timezone.utc),
        )
        db.add(novo)
        inseridos += 1

    for row in parsed_rows:
        upsert(
            row.get("nome"),
            row.get("telefone"),
            row.get("departamento"),
            row.get("sobre_cliente"),
        )

    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail="Não foi possível importar porque existem telefones duplicados ou dados inválidos.",
        ) from exc

    cache_delete_prefix(cache_k("clientes", "list", str(empresa_id)))

    return {
        "ok": True,
        "linhas_lidas": len(parsed_rows),
        "inseridos": inseridos,
        "atualizados": atualizados,
        "ignorados": ignorados,
    }


@router.post("/bulk/colaborador")
def trocar_colaborador_em_massa(
    payload: BulkColaboradorIn,
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.editar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para editar responsáveis dos clientes.",
        )

    if not payload.ids:
        raise HTTPException(status_code=400, detail="Lista de IDs vazia.")

    if payload.colaborador_id is not None:
        ok = (
            db.query(models.Colaborador)
            .filter(models.Colaborador.id == payload.colaborador_id, models.Colaborador.empresa_id == empresa_id)
            .first()
        )
        if not ok:
            raise HTTPException(status_code=400, detail="colaborador_id inválido para esta empresa")

    q = db.query(models.Cliente).filter(
        models.Cliente.empresa_id == empresa_id,
        models.Cliente.id.in_(payload.ids),
    )
    q = _apply_instancias_filter(identity, db, q)

    updated = 0
    for c in q:
        _assert_instancia_acl(identity, db, getattr(c, "instancia_id", None))
        c.colaborador_id = payload.colaborador_id
        db.add(c)
        updated += 1

    db.commit()
    cache_delete_prefix(cache_k("clientes", "list", str(empresa_id)))

    return {"ok": True, "updated": updated}


@router.patch("/{cliente_id}/colaborador")
def trocar_colaborador_unitario(
    cliente_id: int,
    novo_colaborador_id: Optional[int] = Query(None, description="ID do colaborador; null remove"),
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.editar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para editar responsáveis dos clientes.",
        )

    cli = (
        db.query(models.Cliente)
        .filter(models.Cliente.id == cliente_id, models.Cliente.empresa_id == empresa_id)
        .first()
    )
    if not cli:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    _assert_instancia_acl(identity, db, getattr(cli, "instancia_id", None))

    if novo_colaborador_id is not None:
        ok = (
            db.query(models.Colaborador)
            .filter(models.Colaborador.id == novo_colaborador_id, models.Colaborador.empresa_id == empresa_id)
            .first()
        )
        if not ok:
            raise HTTPException(status_code=400, detail="colaborador_id inválido para esta empresa")

    cli.colaborador_id = novo_colaborador_id
    db.add(cli)
    db.commit()
    cache_delete_prefix(cache_k("clientes", "list", str(empresa_id)))
    return {"ok": True}


@router.post("/bulk/departamento")
def trocar_departamento_em_massa(
    payload: BulkDepartamentoIn,
    empresa_id: int = Depends(get_empresa_autorizada),
    db: Session = Depends(get_db),
    identity=Depends(get_current_identity),
):
    perms = _get_perms(identity)
    if "clientes.editar" not in perms:
        raise HTTPException(
            status_code=403,
            detail="Sem permissão para alterar departamentos dos clientes.",
        )

    if not payload.ids:
        raise HTTPException(status_code=400, detail="Lista de IDs vazia.")

    dep_nome: Optional[str] = None
    dep_id = payload.departamento_id

    if dep_id is not None:
        dep_row = (
            db.query(models.Departamento)
            .filter(models.Departamento.id == dep_id, models.Departamento.empresa_id == empresa_id)
            .first()
        )
        if not dep_row:
            raise HTTPException(status_code=400, detail="departamento_id inválido para esta empresa")
        dep_nome = (dep_row.nome or "").strip() or None

    q = db.query(models.Cliente).filter(
        models.Cliente.empresa_id == empresa_id,
        models.Cliente.id.in_(payload.ids),
    )
    q = _apply_instancias_filter(identity, db, q)

    updated = 0
    for c in q:
        _assert_instancia_acl(identity, db, getattr(c, "instancia_id", None))
        c.departamento_id = dep_id
        c.departamento = dep_nome if dep_id is not None else None
        db.add(c)
        updated += 1

    db.commit()
    cache_delete_prefix(cache_k("clientes", "list", str(empresa_id)))

    return {"ok": True, "updated": updated, "departamento_nome": dep_nome}